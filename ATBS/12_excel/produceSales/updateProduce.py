#! python3
# python updateProduce.py
# Your program does the following:
# Loops over all the rows.
# If the row is for garlic, celery, or lemons, changes the price.
# This means your code will need to do the following:
# Open the spreadsheet file.
# For each row, check whether the value in column A is Celery, Garlic, or Lemon.
# If it is, update the price in column B.
# Save the spreadsheet to a new file (so that you don’t lose the old spreadsheet, just in case).

import openpyxl
wb = openpyxl.load_workbook('ProduceSales.xlsx')
sheet = wb['Sheet']

# the produce types and their price
PRICE_UPDATES = {'Garlic': 3.07,
                'Celery': 1.19,
                'Lemon': 1.27}

print(PRICE_UPDATES['Lemon'])

# Loop through the rows and upate the prices
for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
        
wb.save('updatedProduceSales.xlsx')



