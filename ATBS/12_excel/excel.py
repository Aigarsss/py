#! python3
#Chapter 12 – Working with Excel Spreadsheets
# openpyxl needs to be installed

# import openpyxl


# wb = openpyxl.load_workbook('example.xlsx')
# print(type(wb)) # <class 'openpyxl.workbook.workbook.Workbook'>
# print(wb.sheetnames) # book had wb.get_sheet_name() but that is depreciated
# sheet = wb['Sheet1'] # depreciated get_sheet_by_name('Sheet1')
# print(sheet)
# print(type(sheet)) # <class 'openpyxl.worksheet.worksheet.Worksheet'>
# print(sheet.title) # Sheet1
# anotherSheet = wb.active
# print(anotherSheet) #<Worksheet "Sheet1">

#getting cell data
# print(sheet['A1']) # <Cell 'Sheet1'.A1>
# print(sheet['A1'].value) #should return datetime, but doesnt. Returns string
# c = sheet['B1']
# print(c.value)
# print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
# print('Cell ' + c.coordinate + ' is ' + c.value)
# print(sheet['C1'].value)

# print(sheet.cell(row=1, column=2)) # <Cell 'Sheet1'.B1>
# print(sheet.cell(row=1, column=2).value)
# for i in range(1,8,2):
#     print(i, sheet.cell(row=i, column=2).value)

# print(sheet.max_row)        # returns int
# print(sheet.max_column)     # returns int

# get letters and indecies
# from openpyxl.utils import get_column_letter, column_index_from_string # .cell didnt work. utils works
# print(get_column_letter(1))
# print(column_index_from_string('SS'))

# print(tuple(sheet['A1' : 'C3']))
# for rowOfCellObjects in sheet['A1' : 'C3']:
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#     print('--- end of row ---')


# print(sheet[1])
# print(sheet['B'])
# for cellObj in sheet['B']:
#     print(cellObj.value)
'''
Opening sheets

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1'] 
sheet = wb.active
sheet['A1'].column
sheet['A1'].row
sheet['A1'].value
sheet.cell(row=1, column=1).value

create new docs
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet'
print(sheet.title)

creating sheets
we.create_sheet()
we.create_sheet(index=0, title='title')
wb.sheetnames
del wb['<sheet name'] 

wb.save('example_copy.xlsx')

style cells:
from openpyxl.styles import Font
italic24Fon = Font(sz=24, i=True)
sheet['A1'].font = italic24Fon

width and height
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20

merge
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged'
unmerge
sheet.unmerge_cells('A1:D3')

freeze panes
sheet.freeze_panes = 'A2'
sheet.freeze_panes = None

charts
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title = 'First series')
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')



'''

# import census2010 # this was code generated from readCensusExcel.py

# print(census2010.allData['AK']['Anchorage'])
# print(census2010.allData['AK']['Anchorage']['pop'])


##### Creating and saving sheets
# import openpyxl
# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# sheet = wb.active
# print(sheet.title)
# sheet.title = 'Spam Bacon Eggs Sheet'
# print(sheet.title)
# wb.create_sheet()
# wb.create_sheet(index=0, title='The sheet')
# print(wb.sheetnames)
# #wb.remove_sheet(wb['The sheet']) # says this is depreciated
# #wb.remove('The sheet') # says this is depreciated
# del wb['The sheet'] # says this is depreciated
# print(wb.sheetnames)

# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.active
# sheet.title = 'Spam Spam Spam'
# wb.save('example_copy.xlsx')

# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1'] = 'Hello world'
# print(sheet['A1'].value)

# import openpyxl
# from openpyxl.styles import Font

# wb = openpyxl.Workbook()
# sheet = wb.active
# italic24Fon = Font(sz=24, i=True)
# sheet['A1'].font = italic24Fon
# sheet['A1'] = 'Hello World'
# fontObj = Font(name='Times New Roman', b=True)
# sheet['B2'].font = fontObj
# sheet['B2'] = 'Text'
# wb.save('Styled.xlsx')

############## setting formulas
# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1'] = 200
# sheet['A2'] = 300
# sheet['A3'] = '=SUM(A1:A2)'
# wb.save('writeFormula.xlsx')

############# row height and col width
# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1'] = 'Tall row'
# sheet['B2'] = 'Wide column'
# sheet.row_dimensions[1].height = 70
# sheet.column_dimensions['B'].width = 20
# wb.save('dimensions.xlsx')

############# merging

# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.merge_cells('A1:D3')
# sheet['A1'] = 'Twelve cells merged'
# sheet.merge_cells('C5:D5')
# sheet['C5'] = 'Two merged cells'
# wb.save('merged.xlsx')

### Unmerge
# import openpyxl
# wb = openpyxl.load_workbook('merged.xlsx')
# sheet = wb.active
# sheet.unmerge_cells('A1:D3')
# sheet.unmerge_cells('C5:D5')
# wb.save('merged.xlsx')

##### freeze
# import openpyxl
# wb = openpyxl.load_workbook('merged.xlsx')
# sheet = wb.active
# sheet.freeze_panes = 'A2'
# sheet.freeze_panes = None
# wb.save('merged.xlsx')

###### charts

# import openpyxl
# wb = openpyxl.Workbook()
# sheet = wb.active
# for i in range(1, 11): # populate data
#     sheet['A' + str(i)] = i

# refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)

# seriesObj = openpyxl.chart.Series(refObj, title = 'First series')

# chartObj = openpyxl.chart.BarChart()
# chartObj.title = 'My chart'
# chartObj.append(seriesObj)
# sheet.add_chart(chartObj, 'C5')
# wb.save('sampleChart.xlsx')