#! python3
# python multiplicationTable.py <number>

import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

# create workbook and assign sheet
wb = openpyxl.Workbook()
sheet = wb.active

# set up the font
boldFont = Font(b=True)

# assign the headers
for i in range(2, int(sys.argv[1]) + 2):
    # add col headers
    sheet.cell(row = 1, column = i).value = i-1
    sheet.cell(row = 1, column = i).font = boldFont
    # add row headers
    sheet.cell(row = i, column = 1).value = i-1
    sheet.cell(row = i, column = 1).font = boldFont

# generating the values via python
# for row in range(1, sheet.max_row):
#     for col in range (1, sheet.max_column):
#         sheet.cell(row = row + 1, column = col + 1).value = row*col

# the same as above, but generating the values via excel ( '=A2*B1')
for row in range(1, sheet.max_row):
    for col in range (1, sheet.max_column):
        sheet.cell(row = row + 1, column = col + 1).value = '=A' + str(row + 1) + '*' + str(get_column_letter(col + 1)) + str(1)


# create the file
wb.save('multi.xlsx')
