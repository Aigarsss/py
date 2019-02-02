#! python3
# python blankRowInserter.py 3 2 ProduceSales.xlsx
# Create a program blankRowInserter.py that takes two integers and a filename string as 
# command line arguments. Let’s call the first integer N and the second integer M. 
# Starting at row N, the program should insert M blank rows into the spreadsheet. 

# Program works, but only if there are no functions in it. therwise, it wont be correct.

import openpyxl, sys, re
from openpyxl.utils import get_column_letter, column_index_from_string

rowArg = int(sys.argv[1])
intArg = int(sys.argv[2])
fileName = sys.argv[3]

# load the old and the new old spreadsheet
wb = openpyxl.load_workbook(fileName)
nWb = openpyxl.Workbook()
sheet = wb.active
nSheet = nWb.active

# copy the first lines from the doc
lastCopyCell = get_column_letter(sheet.max_column) + str(rowArg - 1)

for rowObj in sheet['A1' : lastCopyCell]:
    for cellObj in rowObj:
        nSheet[cellObj.coordinate] = cellObj.value

# add the blanks and copy the remaining cells (will break functions)
lastCopyCell = get_column_letter(sheet.max_column) + str(sheet.max_row)

for rowObj in sheet['A' + str(rowArg) : lastCopyCell]:
        for cellObj in rowObj:
            #print(cellObj.coordinate, cellObj.value, end=' ')
            row = re.split(r'(\d+)', cellObj.coordinate)
            row[1] = str(int(row[1]) + intArg)
            row = ''.join(row)
            nSheet[row] = cellObj.value
# write the new file
nWb.save('Blanks'+ fileName)