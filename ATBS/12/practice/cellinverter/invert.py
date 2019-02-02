#! python3
# Write a program to invert the row and column of the cells 
# in the spreadsheet. For example, the value at row 5, column 3
# will be at row 3, column 5 (and vice versa). This should be done 
# for all cells in the spreadsheet

import openpyxl, pprint
wb = openpyxl.load_workbook('inverted.xlsx')
nWb = openpyxl.Workbook()
sheet = wb.active
nSheet = nWb.active

# sheetData = {}
sheetData = []
row = []
col = []

# create dictionary that will store the locations
# for i in range(1, sheet.max_row + 1):
#     for n in range(1, sheet.max_column + 1):
#         sheetData[sheet.cell(row = i, column = n).value] = [i,  n]
# pprint.pprint(sheetData)

for i in range(1, sheet.max_row + 1):
    for n in range(1, sheet.max_column + 1):
        row.append(i)
        col.append(n)

sheetData.append(row)
sheetData.append(col)
#pprint.pprint(sheetData)

for i in range(len(sheetData[0])):
    nSheet.cell(row = sheetData[1][i], column = sheetData[0][i]).value = sheet.cell(row = sheetData[0][i], column = sheetData[1][i]).value


#save file
nWb.save('copyInverted.xlsx')
