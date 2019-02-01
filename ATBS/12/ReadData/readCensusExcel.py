#! python3
# python readCensusExcel.py
# Import the openpyxl module.
# Call the openpyxl.load_workbook() function.
# Get a Workbook object.
# Read the active member variable or call the get_sheet_by_name() workbook method.
# Get a Worksheet object.
# Use indexing or the cell() sheet method with row and column keyword arguments.
# Get a Cell object.
# Read the Cell object’s value attribute.

import openpyxl, pprint

print('Opening a workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.active
countyData = {}

# Fill in countyData with each county's population and tracts
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    # each row in the sheet has consensus data for one census tract.
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    # make sure the key for this state exists
    countyData.setdefault(state, {})
    # make sure the key for this county in the state exists.
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})

    # each row represents one census tract, so incerment by one
    countyData[state][county]['tracts'] += 1
    # increase the country pop by the pop in its consensus track
    countyData[state][county]['pop'] += int(pop)

# open a new text file and write the ontents of countyData to it
print('Writing results...')
resultFile =  open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done')


